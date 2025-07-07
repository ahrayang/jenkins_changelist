import os
import re
import logging
import subprocess
from datetime import datetime, timedelta
from collections import defaultdict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def utc_to_kst(utc_str):
    try:
        dt_utc = datetime.strptime(utc_str, "%Y/%m/%d %H:%M:%S")
        dt_kst = dt_utc + timedelta(hours=9)
        return dt_kst.strftime("%Y/%m/%d %H:%M:%S")
    except Exception as e:
        logging.error(f"utc -> kst 변환 실패: {e}")
        return utc_str

def run_command(cmd):
    logging.info(f"[CMD 실행]: {cmd}")
    try:
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True, check=True)
        logging.info(f"[CMD 결과]:\n{result.stdout}")
        return result.stdout
    except subprocess.CalledProcessError as e:
        logging.error(f"[CMD 실행 실패]: {e.stderr}")
        return ""

def get_changes(depot, since, until):
    cmd = f'p4 changes {depot}@{since},{until}'
    logging.info(f"cmd 전송: {cmd}")
    output = run_command(cmd)
    changes = []
    pattern = re.compile(r"^Change\s+(\d+)\s+on\s+(\d{4}/\d{2}/\d{2}(?:\s+\d{2}:\d{2}:\d{2})?)\s+by\s+(\S+)")
    for line in output.splitlines():
        line = line.strip()
        m = re.search(pattern, line)
        if m:
            change_num = m.group(1)
            changes.append(change_num)
    return changes

def parse_describe_grouped(change_num):
    cmd = f"p4 describe -s {change_num}"
    output = run_command(cmd)
    if not output:
        return []
    lines = output.splitlines()
    header_pattern = re.compile(r"^Change\s+\d+\s+by\s+(\S+)@.*\s+on\s+(\d{4}/\d{2}/\d{2}(?:\s+\d{2}:\d{2}:\d{2})?)")
    header_found = False
    description_lines = []
    in_affected = False
    action_to_files = defaultdict(list)
    user, submit_time = "", ""

    for line in lines:
        if not header_found:
            m = re.search(header_pattern, line.strip())
            if m:
                user = m.group(1)
                submit_time = m.group(2)
                if " " not in submit_time:
                    submit_time += " 00:00:00"
                header_found = True
            continue
        if not in_affected:
            if "Affected files" in line:
                in_affected = True
                continue
            else:
                description_lines.append(line.strip())
        else:
            if line.strip() == "":
                continue
            m = re.match(r"^\s*\.\.\.\s+(//\S+)#\d+\s+(\w+)", line)
            if m:
                depot_path = m.group(1)
                action = m.group(2)
                file_name = os.path.basename(depot_path)
                action_to_files[action].append(file_name)

    description = "\n".join(description_lines).strip()
    jira_urls = re.findall(r'(https?://\S+atlassian\.net\S+)', description)
    jira_combined = ", ".join(jira_urls)

    converted = utc_to_kst(submit_time)
    date_part, time_part = converted.split(" ") if " " in converted else (converted, "")

    if user.lower().startswith("jenkins"):
        return []

    rows = []
    for action, files in action_to_files.items():
        rows.append({
            "Change 번호": change_num,
            "날짜": date_part,
            "시간": time_part,
            "작업자": user,
            "설명": description,
            "Action": action,
            "File": ", ".join(files),
            "Jira URL": jira_combined
        })
    return rows

def append_to_excel_with_hyperlink(data_list, excel_file="build_history.xlsx"): # 나중에는 * 처리리
    columns = ["Change 번호", "날짜", "시간", "작업자", "설명", "Action", "File", "Jira URL"]
    df = pd.DataFrame(data_list, columns=columns)
    df.to_excel(excel_file, index=False)

    wb = load_workbook(excel_file)
    ws = wb.active

    jira_col_index = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "Jira URL":
            jira_col_index = idx
            break

    if jira_col_index:
        for row in ws.iter_rows(min_row=2):
            cell = row[jira_col_index - 1]
            if cell.value:
                urls = re.findall(r"https?://\S+atlassian\.net\S*", cell.value)
                if urls:
                    first_url = urls[0]
                    cell.hyperlink = first_url
                    cell.font = Font(color="0000FF", underline="single")

    max_width = 90
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

    wb.save(excel_file)